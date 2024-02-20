"""Parse Excel files for use by the client library."""

from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple, Union
from zipfile import BadZipFile

from geneweaver.core.types import StringOrPath
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.worksheet.worksheet import Worksheet


def get_sheet_names(file_path: StringOrPath) -> List[str]:
    """Retrieve the names of all sheets in an Excel file.

    :param file_path: The path to the Excel file.

    :returns: A list of strings representing the names of all sheets in the file.
    """
    try:
        workbook = load_workbook(filename=file_path, read_only=True)
    except (InvalidFileException, BadZipFile) as e:
        raise ValueError(str(e)) from e
    return workbook.sheetnames


def get_sheet(file_path: StringOrPath, sheet_name: str = None) -> Worksheet:
    """Load a workbook and return a sheet.

    This function opens an Excel workbook and returns the specified sheet if a sheet
    name is provided.

    If no sheet name is provided, it returns the active sheet.

    :param file_path: Path to the Excel (.xlsx) file.
    :param sheet_name: The name of the sheet to open. If not provided, the active sheet
    is returned.

    :returns: The requested worksheet from the Excel workbook.
    """
    workbook = load_workbook(filename=file_path, read_only=True)
    if sheet_name:
        return workbook[sheet_name]
    else:
        return workbook.active


def find_header(
    file_path: StringOrPath, max_rows_to_check: int = 5, sheet_name: str = None
) -> Tuple[bool, int]:
    """Determine if a given Excel worksheet has a header row.

    This function will check up to 'max_rows_to_check' rows from the top to find a
    header row. It assumes that a row can be interpreted as a header if none of its
    values are numeric, if it has more than one column, and if the following row has the
    same number of columns that it does.

    :param file_path: Path to the Excel (.xlsx) file.
    :param max_rows_to_check: The number of rows to check from the top to find a header
    row. Default is 5.
    :param sheet_name: The name of the sheet to open. If not provided, the active sheet
    is returned.

    :returns:
        First index - True if a header row is found, False otherwise.
        Second index - The index of the header row if found, otherwise -1.
    """
    sheet = get_sheet(file_path, sheet_name)
    for i, row in enumerate(sheet.iter_rows(max_row=max_rows_to_check), start=1):
        values = [cell.value for cell in row]
        if (
            len(values) > 1
            and all(value is not None for value in values)
            and all(not isinstance(value, (int, float)) for value in values)
        ):
            next_values = [cell.value for cell in sheet[i + 1]]
            if next_values and len(values) == len(next_values):
                return True, i - 1
    return False, -1


def has_header(
    file_path: str, max_rows_to_check: int = 5, sheet_name: str = None
) -> bool:
    """Summary function to return true/false for if a header is found by `find_header`.

    :param file_path: Path to the Excel (.xlsx) file.
    :param max_rows_to_check: The number of rows to check from the top to find a
    header row.

    :returns:  True if a header row is found, False otherwise.
    """
    return find_header(file_path, max_rows_to_check, sheet_name)[0]


def get_headers(
    file_path: StringOrPath, sheet_name: Optional[str] = None
) -> Tuple[List[str], int]:
    """Read the headers of an Excel (.xlsx) file.

    Read the first row from an Excel (.xlsx) file and return it as a list of strings.

    Assumes that the first row of the Excel file is the header row.

    :param file_path: Path to the Excel file.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: List of column names from the CSV file, and the index of the
    header row.
    """
    headers = []
    file_has_header, header_idx = find_header(file_path, sheet_name=sheet_name)
    if file_has_header:
        headers = read_row(file_path, header_idx, sheet_name=sheet_name)
    return headers, header_idx


def read_row(file_path: str, row_idx: int = 0, sheet_name: str = None) -> List[str]:
    """Get the contents of a row from an Exel (.xlsx) file.

    :param file_path: The file path to the CSV file.
    :param row_idx: The index of the row from which to read the headers. Defaults to 0.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: The contents of a row

    :raises ValueError: If the file is empty or does not contain enough rows.
    """
    sheet = get_sheet(file_path, sheet_name)
    row_idx += 1
    try:
        return [cell.value for cell in sheet[row_idx]]
    except IndexError as e:
        raise ValueError(f"File does not contain a row at index {row_idx}") from e


def read_rows(
    file_path: StringOrPath,
    n_rows: int,
    sheet_name: Optional[str] = None,
    start_row: int = 0,
) -> List[List[str]]:
    """Get the contends of n rows from an Excel (.xlsx) file.

    :param file_path: The file path to the CSV file.
    :param n_rows: The number of rows to read.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.
    :param start_row: The row number to start reading from (0-indexed).

    :returns: The contents of n rows from the Excel file.
    """
    sheet = get_sheet(file_path, sheet_name)
    rows = []
    for row_idx in range(start_row + 1, n_rows):
        try:
            rows.append([cell.value for cell in sheet[row_idx]])
        except IndexError:
            continue
    return rows


def read_to_dict(
    file_path: StringOrPath, start_row: int = 0, sheet_name: Optional[str] = None
) -> List[Dict[str, Union[str, int]]]:
    """Parse an Excel file into a list of dictionaries.

    Parse an Excel file into a list of dictionaries, with keys as column names and
    values as column values.

    :param file_path: The file path to the Excel file.
    :param start_row: The row number to start reading from (0-indexed).
                               This row will be used as the header.
                               Defaults to 0.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: A list of dictionaries, where each dictionary
    represents a row from the Excel file.
    """
    headers = read_row(file_path, start_row, sheet_name=sheet_name)
    sheet = get_sheet(file_path, sheet_name)
    data = []
    for row in sheet.iter_rows(min_row=start_row + 2):
        row_data = {h: cell.value for h, cell in zip(headers, row)}  # noqa: B905
        data.append(row_data)
    return data


def read_to_dict_n_rows(
    file_path: str, n: int, start_row: int = 0, sheet_name: str = None
) -> List[Dict[str, str]]:
    """Parse n lines of an Excel file into a list of dictionaries.

    Parse an Excel file into a list of dictionaries, with keys as column names and
    values as column values. This function will only return the first n rows of data.

    :param file_path: The file path to the Excel file.
    :param n: The number of rows of data to return.
    :param start_row: The row number to start reading from (0-indexed).
                           This row will be used as the header.
                           Defaults to 0.
    :param sheet_name: Name of the sheet to read from. If not provided, the function
    will read from the active sheet.

    :returns: A list of dictionaries, where each dictionary represents
    a row from the Excel file.
    """
    headers = read_row(file_path, start_row, sheet_name=sheet_name)
    sheet = get_sheet(file_path, sheet_name)
    data = []
    for row in sheet.iter_rows(min_row=start_row + 2, max_row=start_row + 1 + n):
        row_data = {h: cell.value for h, cell in zip(headers, row)}  # noqa: B905
        data.append(row_data)
    return data


def default_metadata_cleaner(metadata: str) -> str:
    """Clean metadata from an Excel file.

    :param metadata: The metadata to clean.

    :returns: The cleaned metadata.
    """
    return metadata.replace("\ufeff", "").strip()


def read_metadata(
    file_path: StringOrPath,
    n_rows: int,
    sheet_name: Optional[str] = None,
    start_row: int = 0,
    metadata_cleaner: Optional[Callable[[str], str]] = default_metadata_cleaner,
) -> List[str]:
    """Read the metadata from an Excel file.

    :param file_path: The file path to the CSV or Excel file.
    :param n_rows: The number of rows of metadata to read.
    :param sheet_name: Name of the sheet to read from (for Excel files). If not
    provided, the function will read from the active sheet. Ignored for CSV files.
    :param start_row: The row to start reading from. Defaults to 0.
    :param metadata_cleaner: A function to clean the metadata. Defaults to
    the default_metadata_cleaner function in this namespace.

    :returns: A list of strings representing the metadata from the CSV or Excel file.
    """
    rows = read_rows(file_path, n_rows, sheet_name, start_row)
    return [
        ",".join([metadata_cleaner(str(r)) for r in row if r != "" and r is not None])
        for row in rows
    ]


def get_metadata(
    file_path: Path, sheet: Optional[str] = None
) -> Tuple[List[str], List[List[str]], List[List[str]], List[int], int]:
    """Get the metadata from an Excel file.

    :param file_path: The file path to the Excel file.
    :param sheet: The name of the sheet to read from. If not provided, the function
    will read for all sheets in the file.

    :returns: A tuple containing:
        - A list of sheet names
        - A list of lists of metadata for each sheet
        - A list of lists of headers for each sheet
        - A list of header indices for each sheet
        - The number of sheets in the file
    """
    if sheet:
        sheet_names = [sheet]
        headers, headers_idx = get_headers(file_path, sheet_name=sheet)
        sheet_metadata = [read_metadata(file_path, headers_idx, sheet_name=sheet)]
        headers, headers_idx = [headers], [headers_idx]
        n_sheets = 1
    else:
        sheet_names = get_sheet_names(file_path)
        headers, headers_idx = [], []
        for s in sheet_names:
            h, h_idx = get_headers(file_path, sheet_name=s)
            headers.append(h)
            headers_idx.append(h_idx)

        sheet_metadata = [
            read_metadata(file_path, h, sheet_name=s)
            for s, h in zip(sheet_names, headers_idx)  # noqa: B905
        ]
        n_sheets = len(sheet_names)

    return sheet_names, sheet_metadata, headers, headers_idx, n_sheets
